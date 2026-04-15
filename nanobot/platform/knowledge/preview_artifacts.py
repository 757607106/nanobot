"""Preview artifact generation for office-like source files.

This module intentionally focuses on read-only preview transforms and keeps
indexing/parsing pipelines untouched.
"""

from __future__ import annotations

import base64
import html
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from loguru import logger
from openpyxl import load_workbook

from nanobot.platform.knowledge.models import KnowledgeFile, now_iso
from nanobot.utils.helpers import ensure_dir

_PREVIEW_ARTIFACTS_KEY = "previewArtifacts"
_OFFICE_PREVIEW_KEY = "officePreview"


@dataclass(frozen=True, slots=True)
class GeneratedPreviewArtifact:
    path: Path
    media_type: str
    filename: str
    preview_kind: str


class KnowledgePreviewArtifacts:
    """Build and cache preview artifacts for office documents."""

    _SUPPORTED_OFFICE_SUFFIXES = {".docx", ".xlsx", ".ppt", ".pptx"}

    def __init__(self, *, preview_dir_factory: Callable[[str], Path]) -> None:
        self._preview_dir_factory = preview_dir_factory

    @staticmethod
    def _source_suffix(file: KnowledgeFile) -> str:
        return Path(file.original_filename or file.filename or "").suffix.lower()

    @staticmethod
    def _preview_filename(file: KnowledgeFile, *, preview_kind: str) -> str:
        source_name = str(file.original_filename or file.filename or file.file_id).strip()
        stem = source_name.rsplit(".", 1)[0] if "." in source_name else source_name
        suffix = ".pdf" if preview_kind == "pdf" else ".html"
        return f"{stem}{suffix}"

    @staticmethod
    def _read_cached_entry(file: KnowledgeFile) -> dict[str, Any] | None:
        previews = file.processing_params.get(_PREVIEW_ARTIFACTS_KEY)
        if not isinstance(previews, dict):
            return None
        entry = previews.get(_OFFICE_PREVIEW_KEY)
        if isinstance(entry, dict):
            return entry
        return None

    @staticmethod
    def _build_cached_entry(
        *,
        path: Path,
        file: KnowledgeFile,
        media_type: str,
        preview_kind: str,
    ) -> dict[str, Any]:
        return {
            "status": "ready",
            "path": str(path),
            "mediaType": media_type,
            "previewKind": preview_kind,
            "filename": KnowledgePreviewArtifacts._preview_filename(file, preview_kind=preview_kind),
            "sourceHash": str(file.content_hash or ""),
            "sourcePath": str(file.raw_path or ""),
            "generatedAt": now_iso(),
        }

    @staticmethod
    def _build_failed_entry(
        *,
        file: KnowledgeFile,
        preview_kind: str,
        reason: str,
    ) -> dict[str, Any]:
        normalized_reason = reason.strip() or "preview generation failed"
        status = "unavailable" if "unavailable" in normalized_reason.lower() else "failed"
        return {
            "status": status,
            "path": "",
            "mediaType": "",
            "previewKind": preview_kind,
            "filename": KnowledgePreviewArtifacts._preview_filename(file, preview_kind=preview_kind),
            "sourceHash": str(file.content_hash or ""),
            "sourcePath": str(file.raw_path or ""),
            "generatedAt": now_iso(),
            "reason": normalized_reason,
        }

    @staticmethod
    def _entry_to_artifact(entry: dict[str, Any]) -> GeneratedPreviewArtifact | None:
        path_value = str(entry.get("path") or "").strip()
        if not path_value:
            return None
        path = Path(path_value)
        if not path.exists():
            return None
        media_type = str(entry.get("mediaType") or "").strip() or "text/html"
        preview_kind = str(entry.get("previewKind") or "").strip() or "html"
        filename = str(entry.get("filename") or "").strip() or path.name
        return GeneratedPreviewArtifact(
            path=path,
            media_type=media_type,
            filename=filename,
            preview_kind=preview_kind,
        )

    @staticmethod
    def _xlsx_to_html(path: Path) -> str:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sections: list[str] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            sections.append(f"<section><h2>{html.escape(sheet.title)}</h2><table>")
            for row_idx, row in enumerate(rows):
                cell_tag = "th" if row_idx == 0 else "td"
                sections.append("<tr>")
                for value in row:
                    cell_text = html.escape(str(value or ""))
                    sections.append(f"<{cell_tag}>{cell_text}</{cell_tag}>")
                sections.append("</tr>")
            sections.append("</table></section>")
        body = "\n".join(sections) if sections else "<p>(空工作簿)</p>"
        return (
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>"
            "body{font-family:'SF Pro Text','PingFang SC',sans-serif;padding:24px;line-height:1.6;}"
            "section{margin-bottom:24px;}"
            "table{border-collapse:collapse;width:100%;}"
            "th,td{border:1px solid #d1d5db;padding:8px 10px;vertical-align:top;}"
            "th{background:#f8fafc;text-align:left;}"
            "</style></head><body>"
            f"{body}"
            "</body></html>"
        )

    @staticmethod
    def _docx_to_html(path: Path) -> str:
        try:
            import mammoth
        except ModuleNotFoundError as exc:
            raise RuntimeError("docx preview converter unavailable: mammoth unavailable") from exc

        def _convert_image(image: Any) -> dict[str, str]:
            with image.open() as image_file:
                encoded = base64.b64encode(image_file.read()).decode("ascii")
            content_type = str(getattr(image, "content_type", "") or "").strip() or "image/png"
            return {"src": f"data:{content_type};base64,{encoded}"}

        with path.open("rb") as source:
            result = mammoth.convert_to_html(
                source,
                convert_image=mammoth.images.img_element(_convert_image),
            )
        raw_html = str(result.value or "").strip() or "<p>(空文档)</p>"
        return (
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>"
            "body{font-family:'SF Pro Text','PingFang SC',sans-serif;padding:24px;line-height:1.7;}"
            "table{border-collapse:collapse;max-width:100%;}"
            "th,td{border:1px solid #d1d5db;padding:6px 8px;}"
            "img{max-width:100%;height:auto;}"
            "pre{white-space:pre-wrap;word-break:break-word;}"
            "</style></head><body>"
            f"{raw_html}"
            "</body></html>"
        )

    @staticmethod
    def _convert_presentation_to_pdf(
        source_path: Path,
        preview_dir: Path,
        target_path: Path,
    ) -> tuple[Path | None, str | None]:
        soffice_bin = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice_bin:
            return None, "soffice unavailable"

        command = [
            soffice_bin,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(preview_dir),
            str(source_path),
        ]
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=90,
                check=False,
            )
        except Exception as exc:  # pragma: no cover - defensive
            return None, f"presentation preview converter failed: {exc}"

        if result.returncode != 0:
            stderr = str(result.stderr or "").strip()
            stdout = str(result.stdout or "").strip()
            detail = stderr or stdout or "unknown converter error"
            return None, f"presentation preview converter failed: {detail}"

        generated_path = preview_dir / f"{source_path.stem}.pdf"
        if not generated_path.exists():
            return None, "presentation preview converter did not produce PDF output"

        if generated_path != target_path:
            if target_path.exists():
                target_path.unlink()
            generated_path.replace(target_path)
        return target_path, None

    @staticmethod
    def _merge_preview_entry(file: KnowledgeFile, entry: dict[str, Any]) -> dict[str, Any]:
        processing_params = dict(file.processing_params or {})
        previews = dict(processing_params.get(_PREVIEW_ARTIFACTS_KEY) or {})
        previews[_OFFICE_PREVIEW_KEY] = entry
        processing_params[_PREVIEW_ARTIFACTS_KEY] = previews
        return processing_params

    def resolve_office_preview(
        self,
        kb_id: str,
        file: KnowledgeFile,
    ) -> tuple[GeneratedPreviewArtifact | None, dict[str, Any] | None]:
        suffix = self._source_suffix(file)
        if suffix not in self._SUPPORTED_OFFICE_SUFFIXES:
            return None, None

        raw_path = Path(str(file.raw_path or "").strip())
        if not raw_path.exists():
            return None, None

        cached_entry = self._read_cached_entry(file)
        source_hash = str(file.content_hash or "")
        if (
            cached_entry
            and str(cached_entry.get("sourceHash") or "") == source_hash
            and str(cached_entry.get("sourcePath") or "") == str(raw_path)
        ):
            status = str(cached_entry.get("status") or "").strip().lower()
            if status and status != "ready":
                return None, None
            cached_artifact = self._entry_to_artifact(cached_entry)
            if cached_artifact is not None:
                return cached_artifact, None

        preview_dir = ensure_dir(self._preview_dir_factory(kb_id))
        try:
            if suffix in {".docx", ".xlsx"}:
                preview_path = preview_dir / f"{file.file_id}.office-preview.html"
                preview_html = self._docx_to_html(raw_path) if suffix == ".docx" else self._xlsx_to_html(raw_path)
                preview_path.write_text(preview_html, encoding="utf-8")
                media_type = "text/html"
                preview_kind = "html"
                entry = self._build_cached_entry(
                    path=preview_path,
                    file=file,
                    media_type=media_type,
                    preview_kind=preview_kind,
                )
                artifact = GeneratedPreviewArtifact(
                    path=preview_path,
                    media_type=media_type,
                    filename=self._preview_filename(file, preview_kind=preview_kind),
                    preview_kind=preview_kind,
                )
                return artifact, self._merge_preview_entry(file, entry)

            preview_kind = "pdf"
            preview_path = preview_dir / f"{file.file_id}.office-preview.pdf"
            converted_path, error_message = self._convert_presentation_to_pdf(raw_path, preview_dir, preview_path)
            if converted_path is None:
                logger.debug(
                    "Office preview generation skipped for file_id={} suffix={} reason={}",
                    file.file_id,
                    suffix,
                    error_message,
                )
                failed_entry = self._build_failed_entry(
                    file=file,
                    preview_kind=preview_kind,
                    reason=str(error_message or "presentation preview unavailable"),
                )
                return None, self._merge_preview_entry(file, failed_entry)

            media_type = "application/pdf"
            entry = self._build_cached_entry(
                path=converted_path,
                file=file,
                media_type=media_type,
                preview_kind=preview_kind,
            )
            artifact = GeneratedPreviewArtifact(
                path=converted_path,
                media_type=media_type,
                filename=self._preview_filename(file, preview_kind=preview_kind),
                preview_kind=preview_kind,
            )
            return artifact, self._merge_preview_entry(file, entry)
        except Exception as exc:  # pragma: no cover - safety net
            logger.warning(
                "Office preview generation failed for file_id={} suffix={}: {}",
                file.file_id,
                suffix,
                exc,
            )
            failed_entry = self._build_failed_entry(
                file=file,
                preview_kind="pdf" if suffix in {".ppt", ".pptx"} else "html",
                reason=str(exc),
            )
            return None, self._merge_preview_entry(file, failed_entry)

    @staticmethod
    def collect_preview_paths(file: KnowledgeFile) -> list[str]:
        entry = KnowledgePreviewArtifacts._read_cached_entry(file)
        if not entry:
            return []
        path_value = str(entry.get("path") or "").strip()
        if not path_value:
            return []
        return [path_value]
